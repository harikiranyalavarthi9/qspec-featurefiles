#!/usr/bin/env python3
"""
TestRail Artifact Counts Utility
=================================
Standalone script to fetch and display artifact counts for all projects in TestRail.

Usage:
    1. Install dependencies: pip install requests openpyxl
    2. Update the credentials at the top of this file
    3. Run: python testrail_counts.py
"""

import requests
from requests.auth import HTTPBasicAuth
import json
import logging
from typing import Dict, List, Optional
from urllib.parse import urljoin
from datetime import datetime
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")
    print("   Install with: pip install openpyxl")

# ========================================
# CONFIGURATION - Update these values
# ========================================
TESTRAIL_URL = "https://abc.testrail.io"  # Your TestRail instance URL
USERNAME = "xyz@test.com"              # Your TestRail username/email
API_KEY = "X89gA.8/lBOfRma1rRxV-y5bcPj72X86ELlSNPeFv"                    # Your TestRail API key

# Logging configuration
LOG_LEVEL = logging.INFO  # Change to logging.DEBUG for more detailed logs
LOG_TO_FILE = True        # Set to True to save logs to file
LOG_FILENAME = f"testrail_counts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# Concurrency configuration
MAX_WORKERS = 10  # Maximum number of concurrent API calls

# Rate limiting configuration
MAX_RETRIES = 5           # Maximum number of retry attempts
INITIAL_RETRY_DELAY = 1   # Initial delay in seconds before retry
MAX_RETRY_DELAY = 60      # Maximum delay in seconds between retries
BACKOFF_MULTIPLIER = 2    # Exponential backoff multiplier
REQUEST_DELAY = 0.1       # Small delay between requests (in seconds) to avoid rate limits

# ========================================
# Logging Setup
# ========================================

def setup_logging():
    """Configure logging with console and optional file output"""
    logger = logging.getLogger('testrail_counts')
    logger.setLevel(LOG_LEVEL)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # Console handler with formatting
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(LOG_LEVEL)
    console_formatter = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)
    
    # File handler if enabled
    if LOG_TO_FILE:
        file_handler = logging.FileHandler(LOG_FILENAME, encoding='utf-8')
        file_handler.setLevel(LOG_LEVEL)
        file_formatter = logging.Formatter(
            '%(asctime)s | %(levelname)-8s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(file_formatter)
        logger.addHandler(file_handler)
    
    return logger

# Initialize logger
logger = setup_logging()

# ========================================
# TestRail API Client
# ========================================

class TestRailClient:
    """Simple TestRail API client for fetching artifact counts"""
    
    def __init__(self, base_url: str, username: str, api_key: str):
        self.base_url = base_url.rstrip('/')
        self.auth = HTTPBasicAuth(username, api_key)
        self.session = requests.Session()
        self._lock = threading.Lock()  # Thread-safe logging
        self.rate_limit_hits = 0  # Track rate limit occurrences
        self.total_requests = 0   # Track total API requests
        self.total_retries = 0    # Track total retry attempts
    
    def _build_url(self, endpoint: str) -> str:
        """Build full API URL, handling both direct and index.php? formats"""
        endpoint = endpoint.lstrip('/')
        
        # Try index.php? format first (common for cloud instances)
        if 'index.php?' not in self.base_url and 'api/v2' not in self.base_url:
            return urljoin(self.base_url + '/', f'index.php?/api/v2/{endpoint}')
        
        # Direct /api/v2 format
        if not endpoint.startswith('api/v2'):
            return urljoin(self.base_url + '/', f'api/v2/{endpoint}')
        
        return urljoin(self.base_url + '/', endpoint)
    
    def _get(self, endpoint: str, params: Optional[Dict] = None, retry_count: int = 0) -> Dict:
        """Make GET request to TestRail API with rate limiting and retry logic"""
        url = self._build_url(endpoint)
        
        # Track request count
        with self._lock:
            self.total_requests += 1
        
        # Add small delay to avoid hitting rate limits
        if REQUEST_DELAY > 0:
            time.sleep(REQUEST_DELAY)
        
        try:
            logger.debug(f"GET {endpoint} with params: {params}")
            response = self.session.get(url, auth=self.auth, params=params, timeout=30)
            
            # Handle rate limiting (HTTP 429)
            if response.status_code == 429:
                with self._lock:
                    self.rate_limit_hits += 1
                    self.total_retries += 1
                
                if retry_count < MAX_RETRIES:
                    # Calculate delay with exponential backoff
                    delay = min(INITIAL_RETRY_DELAY * (BACKOFF_MULTIPLIER ** retry_count), MAX_RETRY_DELAY)
                    
                    # Check for Retry-After header
                    retry_after = response.headers.get('Retry-After')
                    if retry_after:
                        try:
                            delay = int(retry_after)
                        except ValueError:
                            pass
                    
                    with self._lock:
                        logger.warning(f"Rate limit hit for {endpoint}. Retrying in {delay}s... (attempt {retry_count + 1}/{MAX_RETRIES})")
                    
                    time.sleep(delay)
                    return self._get(endpoint, params, retry_count + 1)
                else:
                    logger.error(f"Max retries exceeded for {endpoint} due to rate limiting")
                    return {}
            
            # Handle other HTTP errors
            response.raise_for_status()
            data = response.json()
            logger.debug(f"Response from {endpoint}: {len(str(data))} bytes")
            return data
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 503 and retry_count < MAX_RETRIES:
                # Service unavailable - retry with backoff
                with self._lock:
                    self.total_retries += 1
                
                delay = min(INITIAL_RETRY_DELAY * (BACKOFF_MULTIPLIER ** retry_count), MAX_RETRY_DELAY)
                with self._lock:
                    logger.warning(f"Service unavailable for {endpoint}. Retrying in {delay}s... (attempt {retry_count + 1}/{MAX_RETRIES})")
                time.sleep(delay)
                return self._get(endpoint, params, retry_count + 1)
            else:
                logger.error(f"HTTP error calling {endpoint}: {e}")
                return {}
        except requests.exceptions.RequestException as e:
            logger.error(f"Error calling {endpoint}: {e}")
            return {}
    
    def get_stats(self) -> Dict[str, int]:
        """Get API usage statistics"""
        return {
            'total_requests': self.total_requests,
            'rate_limit_hits': self.rate_limit_hits,
            'total_retries': self.total_retries
        }
    
    def get_projects(self) -> List[Dict]:
        """Get all projects"""
        data = self._get('get_projects')
        
        # Handle wrapped response
        if isinstance(data, dict) and 'projects' in data:
            return data['projects']
        elif isinstance(data, list):
            return data
        else:
            return []
    
    def get_project(self, project_id: int) -> Optional[Dict]:
        """Get project details including suite_mode"""
        return self._get(f'get_project/{project_id}')
    
    def count_paginated(self, endpoint: str, params: Optional[Dict] = None) -> int:
        """Count items across all pages using _links.next"""
        total = 0
        current_url = endpoint
        current_params = params or {}
        
        while current_url:
            data = self._get(current_url, current_params)
            
            if not data:
                break
            
            # Add size from current page
            total += data.get('size', 0)
            
            # Check for next page
            next_link = data.get('_links', {}).get('next')
            if next_link:
                # Extract endpoint from next link
                current_url = next_link.split('api/v2/')[-1] if 'api/v2/' in next_link else next_link
                current_params = {}  # Params are in the next URL
            else:
                break
        
        return total
    
    def get_suites(self, project_id: int) -> List[Dict]:
        """Get all suites for a project"""
        data = self._get(f'get_suites/{project_id}')
        
        # Handle wrapped response
        if isinstance(data, dict) and 'suites' in data:
            return data['suites']
        elif isinstance(data, list):
            return data
        else:
            return []
    
    def count_test_cases(self, project_id: int, suite_mode: int) -> int:
        """Count test cases, handling suite_mode = 3 with concurrent API calls"""
        if suite_mode == 3:
            # Multiple suites - need to count per suite concurrently
            suites = self.get_suites(project_id)
            total = 0
            logger.debug(f"    Found {len(suites)} suites for project {project_id}")
            
            def count_suite_cases(suite):
                suite_id = suite.get('id')
                suite_name = suite.get('name')
                count = self.count_paginated(f'get_cases/{project_id}', {'suite_id': suite_id})
                with self._lock:
                    logger.info(f"    └─ Suite '{suite_name}' (ID: {suite_id}): {count} cases")
                return count
            
            # Use ThreadPoolExecutor for concurrent counting
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                future_to_suite = {executor.submit(count_suite_cases, suite): suite for suite in suites}
                for future in as_completed(future_to_suite):
                    try:
                        count = future.result()
                        total += count
                    except Exception as e:
                        suite = future_to_suite[future]
                        logger.error(f"Failed to count cases for suite {suite.get('name')}: {e}")
            
            return total
        else:
            # Single suite or baseline - can query directly
            return self.count_paginated(f'get_cases/{project_id}')
    
    def count_test_results(self, project_id: int) -> int:
        """Count test results by summing results across all runs concurrently"""
        # First, get all runs
        runs_data = self._get(f'get_runs/{project_id}')
        
        if isinstance(runs_data, dict) and 'runs' in runs_data:
            runs = runs_data['runs']
        elif isinstance(runs_data, list):
            runs = runs_data
        else:
            runs = []
        
        if not runs:
            return 0
        
        logger.info(f"    Found {len(runs)} test run(s), fetching results concurrently...")
        
        total_results = 0
        completed_runs = 0
        
        def count_run_results(run):
            run_id = run.get('id')
            run_name = run.get('name', f'Run {run_id}')
            count = self.count_paginated(f'get_results_for_run/{run_id}')
            return run_id, run_name, count
        
        # Use ThreadPoolExecutor for concurrent counting
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_run = {executor.submit(count_run_results, run): run for run in runs}
            for future in as_completed(future_to_run):
                try:
                    run_id, run_name, count = future.result()
                    total_results += count
                    completed_runs += 1
                    
                    with self._lock:
                        logger.info(f"    [{completed_runs}/{len(runs)}] Run '{run_name}' (ID: {run_id}): {count} results")
                except Exception as e:
                    run = future_to_run[future]
                    completed_runs += 1
                    with self._lock:
                        logger.error(f"    [{completed_runs}/{len(runs)}] Failed to count results for run {run.get('id')}: {e}")
        
        logger.info(f"    Total results from all runs: {total_results}")
        return total_results
    
    def get_artifact_counts(self, project_id: int, project_name: str, suite_mode: int) -> Dict[str, int]:
        """Get all artifact counts for a project"""
        logger.info(f"Fetching counts for: {project_name} (ID: {project_id}, Suite Mode: {suite_mode})")
        
        counts = {}
        
        # Milestones (Releases)
        logger.info("Counting milestones...")
        counts['milestones'] = self.count_paginated(f'get_milestones/{project_id}')
        logger.info(f"Milestones: {counts['milestones']}")
        
        # Test Suites (only for suite_mode = 3)
        logger.info("Counting test suites...")
        if suite_mode == 3:
            suites = self.get_suites(project_id)
            counts['test_suites'] = len(suites)
        else:
            counts['test_suites'] = 0  # Not applicable for suite_mode 1 or 2
        logger.info(f"Test Suites: {counts['test_suites']}")
        
        # Test Cases
        logger.info("Counting test cases...")
        counts['test_cases'] = self.count_test_cases(project_id, suite_mode)
        logger.info(f"Test Cases: {counts['test_cases']}")
        
        # Test Plans
        logger.info("Counting test plans...")
        counts['test_plans'] = self.count_paginated(f'get_plans/{project_id}')
        logger.info(f"Test Plans: {counts['test_plans']}")
        
        # Test Runs
        logger.info("Counting test runs...")
        counts['test_runs'] = self.count_paginated(f'get_runs/{project_id}')
        logger.info(f"Test Runs: {counts['test_runs']}")
        
        # Test Results
        logger.info("Counting test results...")
        counts['test_results'] = self.count_test_results(project_id)
        logger.info(f"Test Results: {counts['test_results']}")
        
        logger.info(f"Completed project: {project_name}")
        
        return counts


# ========================================
# Main Execution
# ========================================

def print_summary_table(projects_data: List[Dict]):
    """Print a formatted table of all project counts"""
    print("\n" + "="*140)
    print("TESTRAIL ARTIFACT COUNTS SUMMARY")
    print("="*140)
    
    # Header
    header = f"{'Project Name':<40} | {'ID':<6} | {'Mode':<4} | {'Milestones':<11} | {'Suites':<8} | {'Cases':<8} | {'Plans':<8} | {'Runs':<8} | {'Results':<10}"
    print(header)
    print("-"*140)
    
    # Rows
    for project in projects_data:
        name = project['name'][:38] + '..' if len(project['name']) > 40 else project['name']
        counts = project['counts']
        
        row = (f"{name:<40} | {project['id']:<6} | {project['suite_mode']:<4} | "
               f"{counts['milestones']:<11} | {counts['test_suites']:<8} | {counts['test_cases']:<8} | "
               f"{counts['test_plans']:<8} | {counts['test_runs']:<8} | "
               f"{counts['test_results']:<10}")
        print(row)
    
    print("-"*140)
    
    # Totals
    totals = {
        'milestones': sum(p['counts']['milestones'] for p in projects_data),
        'test_suites': sum(p['counts']['test_suites'] for p in projects_data),
        'test_cases': sum(p['counts']['test_cases'] for p in projects_data),
        'test_plans': sum(p['counts']['test_plans'] for p in projects_data),
        'test_runs': sum(p['counts']['test_runs'] for p in projects_data),
        'test_results': sum(p['counts']['test_results'] for p in projects_data)
    }
    
    total_row = (f"{'TOTAL':<40} | {'':6} | {'':4} | "
                 f"{totals['milestones']:<11} | {totals['test_suites']:<8} | {totals['test_cases']:<8} | "
                 f"{totals['test_plans']:<8} | {totals['test_runs']:<8} | "
                 f"{totals['test_results']:<10}")
    print(total_row)
    print("="*140)


def export_to_json(projects_data: List[Dict], filename: str = "testrail_counts.json"):
    """Export results to JSON file"""
    with open(filename, 'w') as f:
        json.dump(projects_data, f, indent=2)
    logger.info(f"JSON exported to: {filename}")


def export_to_excel(projects_data: List[Dict], filename: str = None):
    """Export results to Excel file with formatting"""
    if not EXCEL_AVAILABLE:
        logger.warning("Excel export skipped - openpyxl not installed")
        return
    
    if filename is None:
        filename = f"testrail_counts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    logger.info(f"Creating Excel workbook: {filename}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestRail Artifact Counts"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Headers
    headers = [
        'Project Name', 
        'Project ID', 
        'Suite Mode', 
        'Milestones', 
        'Test Suites',
        'Test Cases', 
        'Test Plans', 
        'Test Runs', 
        'Test Results'
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Write data rows
    for row_idx, project in enumerate(projects_data, start=2):
        counts = project['counts']
        
        row_data = [
            project['name'],
            project['id'],
            project['suite_mode'],
            counts['milestones'],
            counts['test_suites'],
            counts['test_cases'],
            counts['test_plans'],
            counts['test_runs'],
            counts['test_results']
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Alignment
            if col_idx == 1:  # Project name - left align
                cell.alignment = left_align
            else:  # Numbers - center align
                cell.alignment = center_align
    
    # Calculate totals
    totals_row = len(projects_data) + 2
    totals = {
        'milestones': sum(p['counts']['milestones'] for p in projects_data),
        'test_suites': sum(p['counts']['test_suites'] for p in projects_data),
        'test_cases': sum(p['counts']['test_cases'] for p in projects_data),
        'test_plans': sum(p['counts']['test_plans'] for p in projects_data),
        'test_runs': sum(p['counts']['test_runs'] for p in projects_data),
        'test_results': sum(p['counts']['test_results'] for p in projects_data)
    }
    
    # Write totals row
    totals_data = [
        'TOTAL',
        '',
        '',
        totals['milestones'],
        totals['test_suites'],
        totals['test_cases'],
        totals['test_plans'],
        totals['test_runs'],
        totals['test_results']
    ]
    
    for col_idx, value in enumerate(totals_data, start=1):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.value = value
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        
        if col_idx == 1:
            cell.alignment = left_align
        else:
            cell.alignment = center_align
    
    # Adjust column widths
    column_widths = {
        'A': 40,  # Project Name
        'B': 12,  # Project ID
        'C': 12,  # Suite Mode
        'D': 12,  # Milestones
        'E': 12,  # Test Suites
        'F': 12,  # Test Cases
        'G': 12,  # Test Plans
        'H': 12,  # Test Runs
        'I': 14   # Test Results
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws['A1'] = "Report Generated"
    meta_ws['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta_ws['A2'] = "TestRail URL"
    meta_ws['B2'] = TESTRAIL_URL
    meta_ws['A3'] = "Total Projects"
    meta_ws['B3'] = len(projects_data)
    meta_ws['A4'] = "Username"
    meta_ws['B4'] = USERNAME
    
    # Format metadata
    for row in range(1, 5):
        meta_ws.cell(row=row, column=1).font = Font(bold=True)
    
    meta_ws.column_dimensions['A'].width = 20
    meta_ws.column_dimensions['B'].width = 40
    
    # Save workbook
    wb.save(filename)
    logger.info(f"Excel exported to: {filename}")
    
    return filename


def main():
    """Main execution function"""
    print("="*140)
    print("TESTRAIL ARTIFACT COUNTS UTILITY (with Concurrent API Calls)")
    print("="*140)
    
    # Validate configuration
    if USERNAME == "your-email@example.com" or API_KEY == "your-api-key-here":
        logger.error("ERROR: Please update the credentials at the top of this file!")
        logger.error("   - Set USERNAME to your TestRail email")
        logger.error("   - Set API_KEY to your TestRail API key")
        logger.error("   - Set TESTRAIL_URL to your TestRail instance URL")
        return
    
    logger.info(f"Connecting to: {TESTRAIL_URL}")
    logger.info(f"Username: {USERNAME}")
    
    if LOG_TO_FILE:
        logger.info(f"Logging to file: {LOG_FILENAME}")
    
    # Initialize client
    start_time = datetime.now()
    logger.info(f"Started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    client = TestRailClient(TESTRAIL_URL, USERNAME, API_KEY)
    
    # Get all projects
    logger.info(f"Fetching projects list...")
    projects = client.get_projects()
    
    if not projects:
        logger.error("No projects found or unable to connect!")
        return
    
    logger.info(f"Found {len(projects)} project(s)")
    logger.info("")
    
    # Fetch counts for each project
    projects_data = []
    
    for idx, project in enumerate(projects, start=1):
        project_id = project['id']
        project_name = project['name']
        suite_mode = project.get('suite_mode', 1)
        
        logger.info(f"{'='*100}")
        logger.info(f"Processing Project {idx}/{len(projects)}: {project_name}")
        logger.info(f"{'='*100}")
        
        # Get detailed project info if suite_mode not in list
        if 'suite_mode' not in project:
            logger.debug(f"Fetching project details for {project_name}...")
            project_details = client.get_project(project_id)
            suite_mode = project_details.get('suite_mode', 1) if project_details else 1
        
        try:
            counts = client.get_artifact_counts(project_id, project_name, suite_mode)
            
            projects_data.append({
                'id': project_id,
                'name': project_name,
                'suite_mode': suite_mode,
                'counts': counts
            })
            
            logger.info(f"Successfully completed: {project_name}")
            logger.info("")
            
        except Exception as e:
            logger.error(f"Failed to fetch counts for project {project_name}: {e}", exc_info=True)
            continue
    
    # Calculate elapsed time
    end_time = datetime.now()
    elapsed = end_time - start_time
    logger.info(f"Completed at: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Total time: {elapsed}")
    logger.info("")
    
    # Display API statistics
    stats = client.get_stats()
    logger.info("="*100)
    logger.info("API USAGE STATISTICS")
    logger.info("="*100)
    logger.info(f"Total API Requests: {stats['total_requests']}")
    logger.info(f"Rate Limit Hits: {stats['rate_limit_hits']}")
    logger.info(f"Total Retries: {stats['total_retries']}")
    if stats['total_requests'] > 0:
        rate_limit_percentage = (stats['rate_limit_hits'] / stats['total_requests']) * 100
        logger.info(f"Rate Limit Hit Rate: {rate_limit_percentage:.2f}%")
    logger.info("")
    
    # Display results
    if projects_data:
        print_summary_table(projects_data)
        
        logger.info("")
        logger.info("="*100)
        logger.info("EXPORTING RESULTS")
        logger.info("="*100)
        
        # Export to JSON
        export_to_json(projects_data)
        
        # Export to Excel
        export_to_excel(projects_data)
        
        logger.info("")
        logger.info("="*100)
        logger.info(f"Successfully processed {len(projects_data)}/{len(projects)} project(s)!")
        logger.info("="*100)
        
        if LOG_TO_FILE:
            logger.info(f"Full logs saved to: {LOG_FILENAME}")
    else:
        logger.error("No project data collected!")


if __name__ == "__main__":
    main()

